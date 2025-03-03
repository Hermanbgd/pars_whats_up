import os
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import pickle
import urllib.request
import openpyxl


with open("my_list", "rb") as fp:
    my_list = pickle.load(fp)
# print(my_list)

# phones = ["79608165075", "79277448341"]

count = 1
https://api.whatsapp.com/send/?phone=79899956099&text&type=phone_number&app_absent=0
browser = webdriver.Chrome()
for phone in my_list:
    browser.get(f'https://api.whatsapp.com/send/?phone={phone}&text&type=phone_number&app_absent=0')
    # time.sleep(1)
    # button_1 = browser.find_element(By.LINK_TEXT, "Перейти в чат")
    # button_1.click()
    WebDriverWait(browser, 10).until(EC.element_to_be_clickable((By.LINK_TEXT, "Перейти в чат"))).click()
    # time.sleep(1.5)
    # link_1 = browser.find_element(By.LINK_TEXT, "перейти в WhatsApp Web")
    # link_1.click()
    WebDriverWait(browser, 10).until(EC.element_to_be_clickable((By.LINK_TEXT, "перейти в WhatsApp Web"))).click()
    # time.sleep(25)
    # avatar_button = browser.find_element(By.CSS_SELECTOR, '[title="Profile Details"]')
    # avatar_button.click()
    try:
        WebDriverWait(browser, 30).until(EC.element_to_be_clickable((By.CSS_SELECTOR, '[title="Profile Details"]'))).click()
    # time.sleep(0.5)
    except:
        print(f"{count} профиль не был найден")
        count += 1
        continue
    card = browser.find_element(By.TAG_NAME, "section").find_element(By.CSS_SELECTOR, "*")
    WebDriverWait(browser, 10).until(EC.element_to_be_clickable(card))
    # card_1 = WebDriverWait(browser, 10).until(EC.element_to_be_clickable((By.TAG_NAME, "section")))
    # card = WebDriverWait(browser, 10).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "*")))
    # time.sleep(2)

    try:
        image = card.find_element(By.TAG_NAME, "img").get_attribute("src")
        urllib.request.urlretrieve(image, f"{phone}.png")
    except:
        image = None
    try:
        name = card.find_elements(By.CLASS_NAME, "selectable-text")[1].text
        if name[:2] == "+7":
            name = None
        print(name)
    except:
        name = None

    if image:
        wb = openpyxl.load_workbook("Baz.xlsx")
        ws = wb["Лист1"]
        img = openpyxl.drawing.image.Image(f"{phone}.png")
        img.anchor = f'B{count}'
        img.width = 100
        img.height = 100
        ws.add_image(img)
        wb.save("Baz.xlsx")
        os.remove(f"{phone}.png")
    if name:
        wb = openpyxl.load_workbook("Baz.xlsx")
        ws = wb["Лист1"]
        ws[f"C{count}"] = name
        wb.save("Baz.xlsx")

    count += 1

time.sleep(20)
